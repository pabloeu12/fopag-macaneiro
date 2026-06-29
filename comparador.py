"""
comparador.py
─────────────
Toda a lógica de leitura e comparação das planilhas.
Separado do app visual para facilitar manutenção futura.
"""

import re
import unicodedata
import openpyxl
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configurações ────────────────────────────────────────────
TOLERANCIA = 0.05   # diferença máxima aceita como OK (R$)

# ── Cores do Excel de saída ──────────────────────────────────
COR_VERDE     = PatternFill("solid", fgColor="C6EFCE")
COR_VERMELHO  = PatternFill("solid", fgColor="FFC7CE")
COR_AMARELO   = PatternFill("solid", fgColor="FFEB9C")
COR_CABECALHO = PatternFill("solid", fgColor="1F4E79")
COR_RESUMO    = PatternFill("solid", fgColor="BDD7EE")
FONTE_CAB     = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
FONTE_BOLD    = Font(bold=True, name="Calibri", size=10)
FONTE_NORMAL  = Font(name="Calibri", size=10)

BORDA = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLUNAS_SAIDA = [
    "Matrícula", "Funcionário", "Código(s) Evento", "Nome do Evento",
    "Valor Lançamento", "Referência Sistema", "Provento Sistema",
    "Desconto Sistema", "Tipo Identificado", "Status", "Observação",
]


# ════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES
# ════════════════════════════════════════════════════════════

def limpar(valor) -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    return str(valor).strip().replace("\n", " ").replace("\t", " ")


def normalizar_cabecalho(valor) -> str:
    texto = limpar(valor).lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", texto).strip()


def normalizar_codigo_evento(valor) -> str:
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    texto = limpar(valor)
    match = re.fullmatch(r"(\d+)(?:[,.]0+)?", texto)
    return match.group(1) if match else texto


def para_float(valor) -> float:
    try:
        v = limpar(valor)
        if not v:
            return 0.0
        return float(v.replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def extrair_codigos(nome_coluna: str) -> list[str]:
    """
    Extrai todos os códigos numéricos de um nome de coluna.
    'SALARIO (101)'            → ['101']
    'PENSAO (180)+(327)+(286)' → ['180', '327', '286']
    '50%\\n(115)'               → ['115']
    """
    return re.findall(r"\((\d+)\)", limpar(nome_coluna))


def comparar(val_lanc, ref, prov, desc) -> tuple[str, str]:
    v = para_float(val_lanc)
    if v == 0.0:
        return None, None
    r, p, d = para_float(ref), para_float(prov), para_float(desc)
    if abs(v - r) <= TOLERANCIA and r != 0.0:
        return "OK_REFERENCIA", "Referência"
    if abs(v - p) <= TOLERANCIA and p != 0.0:
        return "OK_PROVENTO", "Provento"
    if abs(v - d) <= TOLERANCIA and d != 0.0:
        return "OK_DESCONTO", "Desconto"
    return "DIVERGENTE", "—"


# ════════════════════════════════════════════════════════════
# LEITURA DAS PLANILHAS
# ════════════════════════════════════════════════════════════

def ler_lancamentos(arquivo) -> tuple[dict, dict, list]:
    """
    Lê a planilha de lançamentos (formato horizontal).
    Retorna:
      lanc_por_mat  → { matricula: { chave_col: float } }
      nomes         → { matricula: nome }
      colunas       → [ { idx, nome, codigos, chave, multi } ]
    """
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb.active
    dados = list(ws.iter_rows(values_only=True))
    wb.close()

    if not dados:
        raise ValueError("Planilha de lançamentos está vazia.")

    cabecalho = dados[0]
    colunas = []
    for idx, nome in enumerate(cabecalho):
        if idx < 2 or nome is None:
            continue
        codigos = extrair_codigos(nome)
        if not codigos:
            continue
        colunas.append({
            "idx":    idx,
            "nome":   limpar(nome),
            "codigos": codigos,
            "chave":  f"{limpar(nome)}||col{idx}",
            "multi":  len(codigos) > 1,
        })

    lanc_por_mat, nomes = {}, {}
    for linha in dados[1:]:
        if not linha or linha[0] is None:
            continue
        try:
            mat = int(float(str(linha[0]).strip()))
        except (ValueError, TypeError):
            continue
        nomes[mat] = limpar(linha[1]) if len(linha) > 1 else ""
        lanc_por_mat[mat] = {
            col["chave"]: para_float(linha[col["idx"]] if col["idx"] < len(linha) else None)
            for col in colunas
        }

    return lanc_por_mat, nomes, colunas


def ler_sistema(arquivo) -> dict:
    """
    Lê a planilha do sistema (formato vertical).
    Retorna:
      sistema → { matricula: { cod_evento: {nome, ref, prov, desc} } }
    """
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb.active
    dados = list(ws.iter_rows(values_only=True))
    wb.close()

    if not dados:
        raise ValueError("Planilha do sistema está vazia.")

    inicio_dados, idxs = _mapear_colunas_sistema(dados)
    sistema = {}
    for linha in dados[inicio_dados:]:
        if not linha or idxs["mat"] >= len(linha) or linha[idxs["mat"]] is None:
            continue
        try:
            mat  = int(float(str(linha[idxs["mat"]]).strip()))
            cod  = normalizar_codigo_evento(linha[idxs["cod"]] if idxs["cod"] < len(linha) else None)
            nome = limpar(linha[idxs["nome"]]) if idxs["nome"] < len(linha) else ""
            ref  = para_float(linha[idxs["ref"]] if idxs["ref"] < len(linha) else None)
            prov = para_float(linha[idxs["prov"]] if idxs["prov"] < len(linha) else None)
            desc = para_float(linha[idxs["desc"]] if idxs["desc"] < len(linha) else None)
        except (ValueError, IndexError, TypeError):
            continue
        if not cod:
            continue
        sistema.setdefault(mat, {})
        if cod in sistema[mat]:
            sistema[mat][cod]["ref"]  += ref
            sistema[mat][cod]["prov"] += prov
            sistema[mat][cod]["desc"] += desc
        else:
            sistema[mat][cod] = {"nome": nome, "ref": ref, "prov": prov, "desc": desc}

    return sistema


def _mapear_colunas_sistema(dados) -> tuple[int, dict]:
    aliases = {
        "mat":  ("matricula",),
        "cod":  ("cod evento", "codigo evento", "cod evento do recibo", "codigo evento do recibo"),
        "nome": ("evento", "nome evento", "nome do evento"),
        "ref":  ("referencia",),
        "prov": ("valor provento", "provento", "provento sistema"),
        "desc": ("valor desconto", "desconto", "desconto sistema"),
    }

    for pos, cabecalho in enumerate(dados):
        normalizados = {
            normalizar_cabecalho(nome): idx
            for idx, nome in enumerate(cabecalho)
            if limpar(nome)
        }
        idxs = {}
        for campo, opcoes in aliases.items():
            idx = next((normalizados[opcao] for opcao in opcoes if opcao in normalizados), None)
            if idx is None:
                break
            idxs[campo] = idx
        else:
            return pos + 1, idxs

    if dados and len(dados[0]) >= 7:
        return 1, {"mat": 0, "cod": 2, "nome": 3, "ref": 4, "prov": 5, "desc": 6}

    raise ValueError("Nao foi possivel identificar as colunas da planilha do sistema.")


# ════════════════════════════════════════════════════════════
# COMPARAÇÃO PRINCIPAL
# ════════════════════════════════════════════════════════════

def executar_comparacao(arquivo_lanc, arquivo_sist) -> list[dict]:
    """
    Orquestra a leitura e comparação completa.
    Retorna lista de dicionários (uma linha por evento).
    """
    lanc_por_mat, nomes, colunas = ler_lancamentos(arquivo_lanc)
    sistema = ler_sistema(arquivo_sist)

    resultados = []

    for mat, valores in lanc_por_mat.items():
        nome_func   = nomes.get(mat, "")
        ev_sistema  = sistema.get(mat)

        for col in colunas:
            val_lanc = valores.get(col["chave"], 0.0)
            if val_lanc == 0.0:
                continue

            codigos    = col["codigos"]
            nome_event = col["nome"]

            # Funcionário não encontrado no sistema
            if ev_sistema is None:
                resultados.append(_linha(
                    mat, nome_func, "+".join(codigos), nome_event,
                    val_lanc, "", "", "", "—",
                    "NAO_ENCONTRADO", "Matrícula não encontrada no sistema",
                ))
                continue

            # Rubrica com múltiplos códigos (ex: PENSÃO)
            if col["multi"]:
                achados    = [c for c in codigos if c in ev_sistema]
                ref_tot    = sum(ev_sistema[c]["ref"]  for c in achados)
                prov_tot   = sum(ev_sistema[c]["prov"] for c in achados)
                desc_tot   = sum(ev_sistema[c]["desc"] for c in achados)

                if not achados:
                    status, tipo, obs = "NAO_ENCONTRADO", "—", "Nenhum dos códigos encontrado"
                else:
                    status, tipo = comparar(val_lanc, ref_tot, prov_tot, desc_tot)
                    if status is None:
                        continue
                    obs = f"Soma de {len(achados)} eventos: {' + '.join(achados)}"

                resultados.append(_linha(
                    mat, nome_func, "+".join(codigos), nome_event,
                    val_lanc, round(ref_tot, 2), round(prov_tot, 2), round(desc_tot, 2),
                    tipo, status, obs,
                ))
                continue

            # Rubrica simples (1 código)
            cod = codigos[0]
            if cod not in ev_sistema:
                resultados.append(_linha(
                    mat, nome_func, cod, nome_event, val_lanc, "", "", "", "—",
                    "NAO_ENCONTRADO", f"Evento {cod} não encontrado no sistema",
                ))
                continue

            ev = ev_sistema[cod]
            status, tipo = comparar(val_lanc, ev["ref"], ev["prov"], ev["desc"])
            if status is None:
                continue

            resultados.append(_linha(
                mat, nome_func, cod, nome_event, val_lanc,
                round(ev["ref"], 2), round(ev["prov"], 2), round(ev["desc"], 2),
                tipo, status, "",
            ))

    return resultados


def _linha(mat, func, cod, evento, val, ref, prov, desc, tipo, status, obs) -> dict:
    return {
        "Matrícula":          mat,
        "Funcionário":        func,
        "Código(s) Evento":   cod,
        "Nome do Evento":     evento,
        "Valor Lançamento":   val,
        "Referência Sistema": ref,
        "Provento Sistema":   prov,
        "Desconto Sistema":   desc,
        "Tipo Identificado":  tipo,
        "Status":             status,
        "Observação":         obs,
    }


# ════════════════════════════════════════════════════════════
# GERAÇÃO DO EXCEL FORMATADO
# ════════════════════════════════════════════════════════════

def gerar_excel(resultados: list[dict]) -> bytes:
    """
    Gera o Excel formatado em memória e retorna os bytes.
    Assim o Streamlit pode oferecer download direto.
    """
    wb = openpyxl.Workbook()

    # ── ABA CONFERÊNCIA ──────────────────────────────────────
    ws = wb.active
    ws.title = "CONFERÊNCIA"

    # Cabeçalho
    for ci, titulo in enumerate(COLUNAS_SAIDA, 1):
        c = ws.cell(row=1, column=ci, value=titulo)
        c.fill = COR_CABECALHO
        c.font = FONTE_CAB
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA
    ws.row_dimensions[1].height = 30

    # Dados
    for ri, reg in enumerate(resultados, 2):
        for ci, col in enumerate(COLUNAS_SAIDA, 1):
            valor = reg.get(col, "")
            c = ws.cell(row=ri, column=ci, value=valor)
            c.font   = FONTE_NORMAL
            c.border = BORDA
            c.alignment = Alignment(vertical="center")
            if col in ("Valor Lançamento", "Referência Sistema",
                       "Provento Sistema", "Desconto Sistema"):
                if isinstance(valor, float):
                    c.number_format = "#,##0.00"

        status = reg.get("Status", "")
        cor = COR_VERDE if "OK" in status else COR_VERMELHO if status == "DIVERGENTE" else COR_AMARELO if status == "NAO_ENCONTRADO" else None
        if cor:
            for ci in range(1, len(COLUNAS_SAIDA) + 1):
                ws.cell(row=ri, column=ci).fill = cor

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "C2"
    larguras = [12, 38, 16, 38, 18, 18, 18, 18, 18, 18, 45]
    for ci, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(ci)].width = larg

    # ── ABA RESUMO ───────────────────────────────────────────
    ws2 = wb.create_sheet("RESUMO")
    total   = len(resultados)
    ok_ref  = sum(1 for r in resultados if r["Status"] == "OK_REFERENCIA")
    ok_prov = sum(1 for r in resultados if r["Status"] == "OK_PROVENTO")
    ok_desc = sum(1 for r in resultados if r["Status"] == "OK_DESCONTO")
    ok_tot  = ok_ref + ok_prov + ok_desc
    diverg  = sum(1 for r in resultados if r["Status"] == "DIVERGENTE")
    nao_enc = sum(1 for r in resultados if r["Status"] == "NAO_ENCONTRADO")
    perc    = round(ok_tot / total * 100, 1) if total else 0

    linhas = [
        ("CONFERÊNCIA DE FOLHA DE PAGAMENTO — MAÇANEIRO", "", ""),
        ("", "", ""),
        ("INDICADOR", "QUANTIDADE", "PERCENTUAL"),
        ("Total comparado",       total,   "100%"),
        ("✅ OK — Referência",    ok_ref,  f"{round(ok_ref/total*100,1) if total else 0}%"),
        ("✅ OK — Provento",      ok_prov, f"{round(ok_prov/total*100,1) if total else 0}%"),
        ("✅ OK — Desconto",      ok_desc, f"{round(ok_desc/total*100,1) if total else 0}%"),
        ("✅ Total OK",           ok_tot,  f"{perc}%"),
        ("❌ Divergentes",        diverg,  f"{round(diverg/total*100,1) if total else 0}%"),
        ("⚠️  Não encontrados",   nao_enc, f"{round(nao_enc/total*100,1) if total else 0}%"),
        ("", "", ""),
        ("PERCENTUAL DE ACERTO GERAL", "", f"{perc}%"),
    ]

    for ri, (a, b, c_val) in enumerate(linhas, 1):
        ws2.cell(row=ri, column=1, value=a)
        ws2.cell(row=ri, column=2, value=b)
        ws2.cell(row=ri, column=3, value=c_val)
        for ci in range(1, 4):
            cel = ws2.cell(row=ri, column=ci)
            cel.border = BORDA
            cel.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            cel.font = FONTE_NORMAL

    ws2.merge_cells("A1:C1")
    ws2["A1"].fill = COR_CABECALHO
    ws2["A1"].font = Font(color="FFFFFF", bold=True, name="Calibri", size=13)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35
    for ci in range(1, 4):
        ws2.cell(row=3, column=ci).fill = COR_RESUMO
        ws2.cell(row=3, column=ci).font = FONTE_BOLD
    for ci in range(1, 4):
        ws2.cell(row=8, column=ci).fill = COR_VERDE
        ws2.cell(row=8, column=ci).font = FONTE_BOLD
        ws2.cell(row=9, column=ci).fill = COR_VERMELHO
        ws2.cell(row=10, column=ci).fill = COR_AMARELO
        ws2.cell(row=12, column=ci).fill = COR_CABECALHO
        ws2.cell(row=12, column=ci).font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
